import openpyxl

file_path = 'Simulation Annuelle Temps Complet 2025 - Modifiée.xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=False) # data_only=False pour lire les formules
    
    max_ifs = 0
    longest_formula = ""
    location = ""
    sheet_found = ""

    print("Recherche de la formule à rallonge...")

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    if_count = formula.upper().count('SI(') + formula.upper().count('IF(')
                    
                    if if_count > max_ifs:
                        max_ifs = if_count
                        longest_formula = formula
                        location = cell.coordinate
                        sheet_found = sheet_name

    print(f"\n--- RÉSULTAT ---")
    print(f"Nombre de SI imbriqués trouvés : {max_ifs}")
    print(f"Localisation : Feuille '{sheet_found}', Cellule {location}")
    print(f"Début de la formule : {longest_formula[:100]}...")
    
except Exception as e:
    print(f"Erreur : {e}")
